from __future__ import annotations

import logging
import os
import secrets
from datetime import date, time
from html import escape
from pathlib import Path
from tempfile import NamedTemporaryFile

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
)
from telegram.constants import ParseMode
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)
from zoneinfo import ZoneInfo

from storage import Storage


logging.basicConfig(
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
    level=logging.INFO,
)
LOGGER = logging.getLogger(__name__)

REMINDER_DAYS = (30, 20, 14, 7, 5, 3, 2, 1)

(
    CONTRACT_TITLE,
    CONTRACT_DESCRIPTION,
    CONTRACT_STAGE_COUNT,
    CONTRACT_STAGE_END_DATE,
    CONTRACT_STAGE_AMOUNT,
    STAGE_CONTRACT_ID,
    STAGE_NAME,
    STAGE_NOTES,
    STAGE_END_DATE,
    STAGE_AMOUNT,
    PAYMENT_CONTRACT_ID,
    PAYMENT_DATE,
    PAYMENT_AMOUNT,
    EDIT_STAGE_NAME,
    EDIT_STAGE_NOTES,
    EDIT_STAGE_END_DATE,
    EDIT_STAGE_AMOUNT,
    EDIT_PAYMENT_DATE,
    EDIT_PAYMENT_AMOUNT,
) = range(19)

MENU_NEW_CONTRACT = "Добавить контракт"
MENU_NEW_STAGE = "Добавить этап"
MENU_NEW_PAYMENT = "Добавить оплату"
MENU_EXPORT_EXCEL = "Выгрузка Excel"
MENU_CONTRACTS = "Показать контракты"
MENU_UPCOMING = "Ближайшие сроки"
MENU_CANCEL = "Отмена"
MENU_INVITE = "Дать доступ"
MENU_ACCESS_LIST = "Список доступов"
CONTRACT_MANAGE_PREFIX = "manage_contract:"
CONTRACT_ADD_STAGE_PREFIX = "add_stage_contract:"
CONTRACT_PAYMENTS_PREFIX = "contract_payments:"
CONTRACT_ADD_PAYMENT_PREFIX = "add_payment_contract:"
STAGE_EDIT_PREFIX = "edit_stage:"
STAGE_DELETE_PREFIX = "delete_stage:"
STAGE_STATUS_PREFIX = "stage_status:"
STAGE_STATUS_SET_PREFIX = "stage_status_set:"
PAYMENT_EDIT_PREFIX = "edit_payment:"
PAYMENT_DELETE_PREFIX = "delete_payment:"
ACCESS_REVOKE_PREFIX = "access_revoke:"

STAGE_STATUSES = {
    "not_started": ("✖", "Не приступили"),
    "in_progress": ("🛠️", "В работе"),
    "waiting_payment": ("🟡", "Ждем оплату"),
    "paid": ("✅", "Оплачен"),
}


def parse_date(raw: str) -> date:
    raw = raw.strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            return date.strptime(raw, fmt)  # type: ignore[attr-defined]
        except AttributeError:
            from datetime import datetime

            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        except ValueError:
            continue
    raise ValueError


def format_due(days_left: int) -> str:
    if days_left < 0:
        return f"просрочено на {abs(days_left)} дн."
    if days_left == 0:
        return "сегодня"
    return f"через {days_left} дн."


def parse_amount(raw: str) -> float:
    normalized = raw.strip().replace(" ", "").replace(",", ".")
    amount = float(normalized)
    if amount < 0:
        raise ValueError
    return round(amount, 2)


def format_date(value: date) -> str:
    return value.strftime("%d-%m-%Y")


def format_amount(amount: float) -> str:
    whole, frac = f"{amount:.2f}".split(".")
    parts = []
    while whole:
        parts.append(whole[-3:])
        whole = whole[:-3]
    return f"{' '.join(reversed(parts))},{frac} ₽"


def format_percent(value: float) -> str:
    return f"{value:.1f}".replace(".", ",") + "%"


def status_emoji(status: str) -> str:
    return STAGE_STATUSES.get(status, STAGE_STATUSES["not_started"])[0]


def status_label(status: str) -> str:
    return STAGE_STATUSES.get(status, STAGE_STATUSES["not_started"])[1]


def safe_sheet_title(title: str, used_titles: set[str]) -> str:
    base = "".join(char for char in title if char not in "[]:*?/\\") or "Контракт"
    base = base[:31]
    candidate = base
    index = 2
    while candidate in used_titles:
        suffix = f" {index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(candidate)
    return candidate


def apply_title(worksheet, cell_range: str, value: str, fill_color: str) -> None:
    worksheet.merge_cells(cell_range)
    cell = worksheet[cell_range.split(":")[0]]
    cell.value = value
    cell.font = Font(size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_header_row(worksheet, row_index: int, fill_color: str) -> None:
    for cell in worksheet[row_index]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_metric_card(label_cell, value_cell, fill_color: str) -> None:
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    label_cell.font = Font(bold=True, color="FFFFFF")
    label_cell.fill = PatternFill("solid", fgColor=fill_color)
    label_cell.alignment = Alignment(horizontal="center")
    label_cell.border = border
    value_cell.font = Font(size=12, bold=True)
    value_cell.fill = PatternFill("solid", fgColor="F8FBFF")
    value_cell.alignment = Alignment(horizontal="center")
    value_cell.border = border


def autosize_worksheet(worksheet) -> None:
    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(column_index)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def build_excel_report(storage: Storage, owner_chat_id: int) -> Path | None:
    contracts = storage.list_contracts(owner_chat_id)
    if not contracts:
        return None

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    contracts_sheet = workbook.create_sheet("Контракты")
    stages_sheet = workbook.create_sheet("Этапы")
    payments_sheet = workbook.create_sheet("Оплаты")

    bold_font = Font(bold=True)
    money_format = "#,##0.00"
    percent_format = "0.0%"
    date_format = "DD-MM-YYYY"
    used_sheet_titles = {"Сводка", "Контракты", "Этапы", "Оплаты"}

    total_contract_amount = 0.0
    total_paid_amount = 0.0
    contract_rows: list[dict] = []
    stage_rows: list[dict] = []
    payment_rows: list[dict] = []

    for contract in contracts:
        stages = storage.list_stages_for_contract(owner_chat_id, contract.id)
        payments = storage.list_payments_for_contract(owner_chat_id, contract.id)
        contract_amount = sum(stage.amount for stage in stages)
        paid_amount = sum(payment.amount for payment in payments)
        debt_amount = max(contract_amount - paid_amount, 0.0)
        paid_ratio = (paid_amount / contract_amount) if contract_amount > 0 else 0.0

        total_contract_amount += contract_amount
        total_paid_amount += paid_amount

        contract_rows.append(
            {
                "contract_id": contract.id,
                "title": contract.title,
                "description": contract.description,
                "deadline": contract.end_date,
                "contract_amount": contract_amount,
                "paid_amount": paid_amount,
                "paid_ratio": paid_ratio,
                "debt_amount": debt_amount,
                "stage_count": len(stages),
                "payment_count": len(payments),
                "stages": stages,
                "payments": payments,
            }
        )

        for stage in stages:
            stage_rows.append(
                {
                    "contract_id": contract.id,
                    "contract_title": contract.title,
                    "stage_name": stage.name,
                    "status": status_label(stage.status),
                    "deadline": stage.end_date,
                    "amount": stage.amount,
                    "notes": stage.notes,
                }
            )

        for payment in payments:
            payment_rows.append(
                {
                    "contract_id": contract.id,
                    "contract_title": contract.title,
                    "payment_date": payment.payment_date,
                    "amount": payment.amount,
                }
            )

    total_debt_amount = max(total_contract_amount - total_paid_amount, 0.0)
    total_paid_ratio = (total_paid_amount / total_contract_amount) if total_contract_amount > 0 else 0.0
    total_debt_ratio = (total_debt_amount / total_contract_amount) if total_contract_amount > 0 else 0.0

    apply_title(summary_sheet, "A1:F1", "Сводка по контрактам", "1F4E78")
    summary_rows = [
        ("Дата выгрузки", date.today()),
        ("Количество контрактов", len(contracts)),
        ("Общая сумма контрактов", total_contract_amount),
        ("Оплачено всего", total_paid_amount),
        ("Оплачено, %", total_paid_ratio),
        ("Долг всего", total_debt_amount),
        ("Долг, %", total_debt_ratio),
    ]
    for index, (label, value) in enumerate(summary_rows, start=3):
        summary_sheet.cell(row=index, column=1, value=label).font = bold_font
        summary_sheet.cell(row=index, column=2, value=value)
        summary_sheet.cell(row=index, column=1).fill = PatternFill("solid", fgColor="DCE6F1")
    summary_sheet["D3"] = "Оплачено"
    summary_sheet["D4"] = total_paid_amount
    summary_sheet["E3"] = "Оплачено, %"
    summary_sheet["E4"] = total_paid_ratio
    summary_sheet["F3"] = "Долг"
    summary_sheet["F4"] = total_debt_amount
    summary_sheet["G3"] = "Долг, %"
    summary_sheet["G4"] = total_debt_ratio
    for cell_ref, fill_color in (("D3", "2E8B57"), ("E3", "2E8B57"), ("F3", "C0504D"), ("G3", "C0504D")):
        summary_sheet[cell_ref].font = Font(bold=True, color="FFFFFF")
        summary_sheet[cell_ref].fill = PatternFill("solid", fgColor=fill_color)
        summary_sheet[cell_ref].alignment = Alignment(horizontal="center")
    for cell_ref in ("D4", "E4", "F4", "G4"):
        summary_sheet[cell_ref].font = Font(size=12, bold=True)
        summary_sheet[cell_ref].alignment = Alignment(horizontal="center")
        summary_sheet[cell_ref].fill = PatternFill("solid", fgColor="F8FBFF")
    for row_index in (5, 6, 8):
        summary_sheet.cell(row=row_index, column=2).number_format = money_format
    summary_sheet.cell(row=3, column=2).number_format = date_format
    for row_index in (7, 9):
        summary_sheet.cell(row=row_index, column=2).number_format = percent_format
    summary_sheet["D4"].number_format = money_format
    summary_sheet["E4"].number_format = percent_format
    summary_sheet["F4"].number_format = money_format
    summary_sheet["G4"].number_format = percent_format

    summary_sheet["I3"] = "Статус"
    summary_sheet["J3"] = "Сумма"
    summary_sheet["I4"] = "Оплачено"
    summary_sheet["I5"] = "Долг"
    summary_sheet["J4"] = total_paid_amount
    summary_sheet["J5"] = total_debt_amount
    summary_sheet["J4"].number_format = money_format
    summary_sheet["J5"].number_format = money_format
    chart = DoughnutChart()
    chart.title = "Оплата vs долг"
    chart.holeSize = 58
    labels = Reference(summary_sheet, min_col=9, min_row=4, max_row=5)
    data = Reference(summary_sheet, min_col=10, min_row=3, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.height = 7
    chart.width = 9
    summary_sheet.add_chart(chart, "I7")

    apply_title(contracts_sheet, "A1:J1", "Сводная таблица контрактов", "264653")
    contracts_sheet.append(
        [
            "ID контракта",
            "Контракт",
            "Описание",
            "Общий дедлайн",
            "Сумма контракта",
            "Оплачено",
            "Оплачено, %",
            "Долг",
            "Кол-во этапов",
            "Кол-во оплат",
        ]
    )
    style_header_row(contracts_sheet, 2, "264653")
    for row in contract_rows:
        contracts_sheet.append(
            [
                row["contract_id"],
                row["title"],
                row["description"],
                row["deadline"],
                row["contract_amount"],
                row["paid_amount"],
                row["paid_ratio"],
                row["debt_amount"],
                row["stage_count"],
                row["payment_count"],
            ]
        )
    for row in contracts_sheet.iter_rows(min_row=3, max_row=contracts_sheet.max_row):
        row[4].number_format = money_format
        row[5].number_format = money_format
        row[6].number_format = percent_format
        row[7].number_format = money_format

    apply_title(stages_sheet, "A1:G1", "Все этапы", "8A5A44")
    stages_sheet.append(
        [
            "ID контракта",
            "Контракт",
            "Этап",
            "Статус",
            "Дедлайн этапа",
            "Сумма этапа",
            "Примечание",
        ]
    )
    style_header_row(stages_sheet, 2, "8A5A44")
    for row in stage_rows:
        stages_sheet.append(
            [
                row["contract_id"],
                row["contract_title"],
                row["stage_name"],
                row["status"],
                row["deadline"],
                row["amount"],
                row["notes"],
            ]
        )
    for row in stages_sheet.iter_rows(min_row=3, max_row=stages_sheet.max_row):
        row[5].number_format = money_format

    apply_title(payments_sheet, "A1:D1", "Все оплаты", "2A9D8F")
    payments_sheet.append(
        [
            "ID контракта",
            "Контракт",
            "Дата оплаты",
            "Сумма оплаты",
        ]
    )
    style_header_row(payments_sheet, 2, "2A9D8F")
    for row in payment_rows:
        payments_sheet.append(
            [
                row["contract_id"],
                row["contract_title"],
                row["payment_date"],
                row["amount"],
            ]
        )
    for row in payments_sheet.iter_rows(min_row=3, max_row=payments_sheet.max_row):
        row[3].number_format = money_format

    for contract_row in contract_rows:
        sheet_title = safe_sheet_title(contract_row["title"], used_sheet_titles)
        contract_sheet = workbook.create_sheet(sheet_title)
        apply_title(contract_sheet, "A1:H1", f"Дашборд: {contract_row['title']}", "0B6E4F")
        contract_sheet["A2"] = "Описание"
        contract_sheet["B2"] = contract_row["description"] or "-"
        contract_sheet["E2"] = "Дедлайн"
        contract_sheet["F2"] = contract_row["deadline"]
        contract_sheet["F2"].number_format = date_format
        contract_sheet["A4"] = "Сумма контракта"
        contract_sheet["B4"] = contract_row["contract_amount"]
        contract_sheet["C4"] = "Оплачено"
        contract_sheet["D4"] = contract_row["paid_amount"]
        contract_sheet["E4"] = "Оплачено, %"
        contract_sheet["F4"] = contract_row["paid_ratio"]
        contract_sheet["G4"] = "Долг"
        contract_sheet["H4"] = contract_row["debt_amount"]
        contract_sheet["I4"] = "Этапов"
        contract_sheet["J4"] = contract_row["stage_count"]
        contract_sheet["K4"] = "Оплат"
        contract_sheet["L4"] = contract_row["payment_count"]
        for pair in (("A4", "B4"), ("C4", "D4"), ("E4", "F4"), ("G4", "H4"), ("I4", "J4"), ("K4", "L4")):
            style_metric_card(contract_sheet[pair[0]], contract_sheet[pair[1]], "0B6E4F")
        contract_sheet["B4"].number_format = money_format
        contract_sheet["D4"].number_format = money_format
        contract_sheet["F4"].number_format = percent_format
        contract_sheet["H4"].number_format = money_format

        contract_sheet["N2"] = "Статус"
        contract_sheet["O2"] = "Сумма"
        contract_sheet["N3"] = "Оплачено"
        contract_sheet["N4"] = "Долг"
        contract_sheet["O3"] = contract_row["paid_amount"]
        contract_sheet["O4"] = contract_row["debt_amount"]
        contract_sheet["O3"].number_format = money_format
        contract_sheet["O4"].number_format = money_format
        contract_chart = DoughnutChart()
        contract_chart.title = "Оплата по контракту"
        contract_chart.holeSize = 62
        chart_labels = Reference(contract_sheet, min_col=14, min_row=3, max_row=4)
        chart_data = Reference(contract_sheet, min_col=15, min_row=2, max_row=4)
        contract_chart.add_data(chart_data, titles_from_data=True)
        contract_chart.set_categories(chart_labels)
        contract_chart.height = 6
        contract_chart.width = 8
        contract_sheet.add_chart(contract_chart, "N6")

        apply_title(contract_sheet, "A7:G7", "Этапы контракта", "8A5A44")
        contract_sheet.append(
            ["Этап", "Статус", "Дедлайн", "Сумма", "Примечание", "", ""]
        )
        style_header_row(contract_sheet, 8, "8A5A44")
        for stage in contract_row["stages"]:
            contract_sheet.append(
                [
                    stage.name,
                    status_label(stage.status),
                    stage.end_date,
                    stage.amount,
                    stage.notes,
                    "",
                    "",
                ]
            )
        stage_start_row = 9
        for row in contract_sheet.iter_rows(min_row=stage_start_row, max_row=contract_sheet.max_row):
            if row[0].value is None:
                break
            row[2].number_format = date_format
            row[3].number_format = money_format

        payments_title_row = contract_sheet.max_row + 2
        apply_title(contract_sheet, f"A{payments_title_row}:D{payments_title_row}", "Оплаты по контракту", "2A9D8F")
        payment_header_row = payments_title_row + 1
        contract_sheet.append(["Дата оплаты", "Сумма", "", ""])
        style_header_row(contract_sheet, payment_header_row, "2A9D8F")
        for payment in contract_row["payments"]:
            contract_sheet.append([payment.payment_date, payment.amount, "", ""])
        for row in contract_sheet.iter_rows(min_row=payment_header_row + 1, max_row=contract_sheet.max_row):
            if row[0].value is None:
                break
            row[0].number_format = date_format
            row[1].number_format = money_format

    for worksheet in (summary_sheet, contracts_sheet, stages_sheet, payments_sheet):
        autosize_worksheet(worksheet)
        worksheet.freeze_panes = "A3"
    for worksheet in workbook.worksheets[4:]:
        autosize_worksheet(worksheet)
        worksheet.freeze_panes = "A8"

    report_file = NamedTemporaryFile(prefix="contracts_report_", suffix=".xlsx", delete=False)
    report_path = Path(report_file.name)
    report_file.close()
    workbook.save(report_path)
    return report_path


def render_contract(contract) -> str:
    description = f"\nОписание: {escape(contract.description)}" if contract.description else ""
    return (
        f"#{contract.id} <b>{escape(contract.title)}</b>\n"
        f"Срок окончания: {format_date(contract.end_date)}{description}"
    )


def render_contract_with_stages(contract, stages) -> str:
    total_amount = sum(stage.amount for stage in stages)
    paid_amount = getattr(contract, "paid_amount", 0.0)
    debt_amount = max(total_amount - paid_amount, 0.0)
    paid_percent = (paid_amount / total_amount * 100) if total_amount > 0 else 0.0
    lines = [
        f"Контракт #{contract.id}: <b>{escape(contract.title)}</b>",
        f"Общий дедлайн: {format_date(contract.end_date)}",
        f"Общая сумма: {format_amount(total_amount)}",
        f"Оплачено: {format_amount(paid_amount)} из {format_amount(total_amount)} ({format_percent(paid_percent)})",
        f"Долг: {format_amount(debt_amount)}",
    ]
    if contract.description:
        lines.append(f"Описание: {escape(contract.description)}")
    if stages:
        lines.append("")
        for stage in stages:
            lines.append(
                f"{status_emoji(stage.status)} Этап {stage.position} - {format_amount(stage.amount)} - {format_date(stage.end_date)}"
            )
    else:
        lines.append("")
        lines.append("Этапов пока нет.")
    return "\n".join(lines)


def render_stage(stage) -> str:
    notes = f"\nПримечание: {escape(stage.notes)}" if stage.notes else ""
    return (
        f"{status_emoji(stage.status)} <b>{escape(stage.name)}</b>\n"
        f"Контракт: {escape(stage.contract_title)}\n"
        f"Статус: {status_label(stage.status)}\n"
        f"Срок окончания: {format_date(stage.end_date)}\n"
        f"Сумма: {format_amount(stage.amount)}{notes}"
    )


def main_menu_markup() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [MENU_NEW_CONTRACT, MENU_NEW_STAGE],
            [MENU_NEW_PAYMENT, MENU_CONTRACTS],
            [MENU_UPCOMING, MENU_EXPORT_EXCEL],
            [MENU_INVITE, MENU_ACCESS_LIST],
        ],
        resize_keyboard=True,
    )


def contract_selection_markup(contracts) -> ReplyKeyboardMarkup:
    keyboard = [[KeyboardButton(f"{contract.id} | {contract.title}")] for contract in contracts]
    keyboard.append([KeyboardButton(MENU_CANCEL)])
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)


def cancel_markup() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([[KeyboardButton(MENU_CANCEL)]], resize_keyboard=True, one_time_keyboard=True)


def is_menu_text(raw: str) -> bool:
    return raw in {
        MENU_NEW_CONTRACT,
        MENU_NEW_STAGE,
        MENU_NEW_PAYMENT,
        MENU_EXPORT_EXCEL,
        MENU_CONTRACTS,
        MENU_UPCOMING,
        MENU_INVITE,
        MENU_ACCESS_LIST,
        MENU_CANCEL,
    }


def viewer_menu_markup() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [MENU_CONTRACTS, MENU_UPCOMING],
            [MENU_EXPORT_EXCEL],
        ],
        resize_keyboard=True,
    )


def payment_list_markup(contract_id: int, payments, can_edit: bool) -> InlineKeyboardMarkup | None:
    rows = []
    if can_edit:
        rows.append([InlineKeyboardButton("Добавить оплату", callback_data=f"{CONTRACT_ADD_PAYMENT_PREFIX}{contract_id}")])
        for payment in payments:
            rows.append(
                [
                    InlineKeyboardButton(
                        f"Изменить {format_date(payment.payment_date)} · {format_amount(payment.amount)}",
                        callback_data=f"{PAYMENT_EDIT_PREFIX}{payment.id}",
                    ),
                    InlineKeyboardButton("Удалить", callback_data=f"{PAYMENT_DELETE_PREFIX}{payment.id}"),
                ]
            )
    return InlineKeyboardMarkup(rows) if rows else None


def resolve_data_scope(update: Update, context: ContextTypes.DEFAULT_TYPE) -> tuple[int, bool] | None:
    storage: Storage = context.application.bot_data["storage"]
    chat = update.effective_chat
    user = update.effective_user
    if chat is None or user is None:
        return None
    if chat.type != "private":
        return None
    owner_chat_id = storage.get_shared_owner(user.id)
    if owner_chat_id is not None and owner_chat_id != chat.id:
        return owner_chat_id, False
    return chat.id, True


async def require_owner(update: Update, context: ContextTypes.DEFAULT_TYPE) -> tuple[int, bool] | None:
    scope = resolve_data_scope(update, context)
    if scope is None:
        if update.effective_message is not None:
            await update.effective_message.reply_text("Используйте бота в личном чате.")
        return None
    owner_chat_id, can_edit = scope
    if not can_edit:
        if update.effective_message is not None:
            await update.effective_message.reply_text(
                "У вас есть только просмотр. Изменять данные может только владелец.",
                reply_markup=viewer_menu_markup(),
            )
        return None
    return owner_chat_id, can_edit


def stage_actions_markup(stage_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("Изменить", callback_data=f"{STAGE_EDIT_PREFIX}{stage_id}"),
                InlineKeyboardButton("Удалить", callback_data=f"{STAGE_DELETE_PREFIX}{stage_id}"),
                InlineKeyboardButton("Статус", callback_data=f"{STAGE_STATUS_PREFIX}{stage_id}"),
            ]
        ]
    )


def stage_status_markup(stage_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("✖ Не приступили", callback_data=f"{STAGE_STATUS_SET_PREFIX}{stage_id}:not_started")],
            [InlineKeyboardButton("🛠️ В работе", callback_data=f"{STAGE_STATUS_SET_PREFIX}{stage_id}:in_progress")],
            [InlineKeyboardButton("🟡 Ждем оплату", callback_data=f"{STAGE_STATUS_SET_PREFIX}{stage_id}:waiting_payment")],
            [InlineKeyboardButton("✅ Оплачен", callback_data=f"{STAGE_STATUS_SET_PREFIX}{stage_id}:paid")],
        ]
    )


def contract_actions_markup(contract_id: int, can_edit: bool) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton("Показать список платежей", callback_data=f"{CONTRACT_PAYMENTS_PREFIX}{contract_id}")]]
    if can_edit:
        rows.insert(0, [InlineKeyboardButton("Управлять этапами", callback_data=f"{CONTRACT_MANAGE_PREFIX}{contract_id}")])
    return InlineKeyboardMarkup(rows)


def contract_manage_actions_markup(contract_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton("Добавить этап", callback_data=f"{CONTRACT_ADD_STAGE_PREFIX}{contract_id}")]]
    )


def revoke_access_markup(viewer_user_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton("Отозвать доступ", callback_data=f"{ACCESS_REVOKE_PREFIX}{viewer_user_id}")]]
    )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    storage: Storage = context.application.bot_data["storage"]
    chat = update.effective_chat
    user = update.effective_user
    if chat is None or update.message is None or user is None:
        return
    storage.register_chat(chat.id)
    if context.args and chat.type == "private":
        token_arg = context.args[0]
        if token_arg.startswith("view_"):
            owner_chat_id = storage.consume_invite_token(
                token_arg.removeprefix("view_"),
                user.id,
                user.username or "",
                user.full_name or "",
            )
            if owner_chat_id is not None:
                await update.message.reply_text(
                    "Доступ к контрактам выдан. Теперь вы можете смотреть список и ближайшие сроки.",
                    reply_markup=viewer_menu_markup(),
                )
                return
            await update.message.reply_text("Ссылка недействительна или уже использована.", reply_markup=viewer_menu_markup())
            return

    scope = resolve_data_scope(update, context)
    if scope is None:
        await update.message.reply_text("Используйте бота в личном чате.")
        return
    _, can_edit = scope
    await update.message.reply_text(
        "Бот для контрактов готов.\n\n"
        "Команды:\n"
        "/new_contract - добавить контракт\n"
        "/new_stage - добавить этап к контракту\n"
        "/new_payment - добавить оплату по контракту\n"
        "/export_excel - выгрузить Excel-отчет\n"
        "/contracts - список контрактов и этапов\n"
        "/upcoming - ближайшие дедлайны\n"
        "/invite_viewer - дать ссылку на просмотр\n"
        "/delete_contract ID - удалить контракт\n"
        "/delete_stage ID - удалить этап\n"
        "/help - подсказка",
        reply_markup=main_menu_markup() if can_edit else viewer_menu_markup(),
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return
    scope = resolve_data_scope(update, context)
    can_edit = bool(scope and scope[1])
    await update.message.reply_text(
        "Формат дат: DD-MM-YYYY\n"
        "Формат суммы: 150000 или 150000.50\n"
        "Напоминания отправляются автоматически за 30, 20, 14, 7, 5, 3, 2 и 1 день до окончания контракта или этапа.\n\n"
        "Типовой сценарий:\n"
        "1. /new_contract\n"
        "2. Бот сам спросит этапы, суммы и сроки\n"
        "3. /new_payment когда пришла оплата\n"
        "4. /export_excel чтобы получить сводку в Excel\n"
        "5. /upcoming",
        reply_markup=main_menu_markup() if can_edit else viewer_menu_markup(),
    )


async def invite_viewer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None:
        return
    owner_chat_id, _ = owner_scope
    token = secrets.token_urlsafe(18)
    storage.create_invite_token(owner_chat_id, token)
    bot_info = await context.bot.get_me()
    invite_link = f"https://t.me/{bot_info.username}?start=view_{token}"
    await update.message.reply_text(
        "Отправьте эту ссылку человеку, которому хотите дать просмотр:\n"
        f"{invite_link}",
        reply_markup=main_menu_markup(),
    )


async def access_list(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None:
        return
    owner_chat_id, _ = owner_scope
    grants = storage.viewer_grants_for_owner(owner_chat_id)
    if not grants:
        await update.message.reply_text("Доступы пока никому не выданы.", reply_markup=main_menu_markup())
        return
    for grant in grants:
        username = grant["viewer_username"] or ""
        name = grant["viewer_name"] or ""
        if not username or not name:
            try:
                chat_info = await context.bot.get_chat(grant["viewer_user_id"])
                username = chat_info.username or username
                name = chat_info.full_name or name
                storage.update_viewer_profile(
                    owner_chat_id,
                    grant["viewer_user_id"],
                    username or "",
                    name or "",
                )
            except Exception:
                LOGGER.exception("Failed to refresh viewer profile for user_id=%s", grant["viewer_user_id"])
        username_line = f"@{username}" if username else "username не указан"
        name_line = name or "Имя не указано"
        await update.message.reply_text(
            f"{name_line}\n{username_line}\nID: {grant['viewer_user_id']}",
            reply_markup=revoke_access_markup(grant["viewer_user_id"]),
        )
    await update.message.reply_text("Список доступов показан.", reply_markup=main_menu_markup())


async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    storage: Storage = context.application.bot_data["storage"]
    scope = resolve_data_scope(update, context)
    if scope is None or update.message is None:
        return
    owner_chat_id, can_edit = scope
    report_path = build_excel_report(storage, owner_chat_id)
    if report_path is None:
        await update.message.reply_text(
            "Контрактов пока нет, выгружать нечего.",
            reply_markup=main_menu_markup() if can_edit else viewer_menu_markup(),
        )
        return
    try:
        with report_path.open("rb") as report_file:
            await update.message.reply_document(
                document=report_file,
                filename=f"contracts_report_{date.today().strftime('%d-%m-%Y')}.xlsx",
                caption="Выгрузка по контрактам готова.",
                reply_markup=main_menu_markup() if can_edit else viewer_menu_markup(),
            )
    finally:
        try:
            report_path.unlink()
        except FileNotFoundError:
            pass


async def revoke_access(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    query = update.callback_query
    if owner_scope is None or query is None or query.data is None:
        return
    owner_chat_id, _ = owner_scope
    await query.answer()
    try:
        viewer_user_id = int(query.data.split(":", 1)[1])
    except (IndexError, ValueError):
        await query.edit_message_reply_markup(reply_markup=None)
        return
    revoked = storage.revoke_viewer_access(owner_chat_id, viewer_user_id)
    await query.edit_message_reply_markup(reply_markup=None)
    await query.message.reply_text(
        "Доступ отозван." if revoked else "Доступ не найден.",
        reply_markup=main_menu_markup(),
    )


async def new_contract(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None:
        return ConversationHandler.END
    context.user_data.clear()
    await update.message.reply_text("Введите название контракта.", reply_markup=cancel_markup())
    return CONTRACT_TITLE


async def contract_title(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message is None or not update.message.text:
        return CONTRACT_TITLE
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Сначала введите название контракта или нажмите Отмена.", reply_markup=cancel_markup())
        return CONTRACT_TITLE
    context.user_data["contract_title"] = raw
    await update.message.reply_text("Введите описание контракта или отправьте '-' если не нужно.", reply_markup=cancel_markup())
    return CONTRACT_DESCRIPTION


async def contract_description(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message is None or not update.message.text:
        return CONTRACT_DESCRIPTION
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Сначала введите описание или '-' либо нажмите Отмена.", reply_markup=cancel_markup())
        return CONTRACT_DESCRIPTION
    context.user_data["contract_description"] = "" if raw == "-" else raw
    await update.message.reply_text("Сколько этапов в контракте? Введите число.", reply_markup=cancel_markup())
    return CONTRACT_STAGE_COUNT


async def contract_stage_count(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message is None or not update.message.text or update.effective_chat is None:
        return CONTRACT_STAGE_COUNT
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Введите количество этапов числом или нажмите Отмена.", reply_markup=cancel_markup())
        return CONTRACT_STAGE_COUNT
    try:
        stage_count = int(raw)
    except ValueError:
        await update.message.reply_text("Количество этапов должно быть числом.")
        return CONTRACT_STAGE_COUNT
    if stage_count <= 0:
        await update.message.reply_text("Количество этапов должно быть больше нуля.")
        return CONTRACT_STAGE_COUNT

    context.user_data["contract_stage_count"] = stage_count
    context.user_data["contract_stage_index"] = 1
    context.user_data["contract_stage_items"] = []
    await update.message.reply_text(
        "Этап 1.\nВведите дедлайн этапа в формате DD-MM-YYYY.",
        reply_markup=cancel_markup(),
    )
    return CONTRACT_STAGE_END_DATE


async def contract_stage_end_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message is None or not update.message.text:
        return CONTRACT_STAGE_END_DATE
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Введите дату этапа в формате YYYY-MM-DD или нажмите Отмена.", reply_markup=cancel_markup())
        return CONTRACT_STAGE_END_DATE
    try:
        end_date = parse_date(raw)
    except ValueError:
        await update.message.reply_text("Не удалось распознать дату. Нужен формат YYYY-MM-DD.")
        return CONTRACT_STAGE_END_DATE

    items = context.user_data["contract_stage_items"]
    if items and end_date < items[-1]["end_date"]:
        await update.message.reply_text(
            f"Дата этапа не может быть раньше предыдущего этапа ({items[-1]['end_date'].isoformat()}).",
            reply_markup=cancel_markup(),
        )
        return CONTRACT_STAGE_END_DATE

    context.user_data["pending_contract_stage_end_date"] = end_date
    stage_index = context.user_data["contract_stage_index"]
    await update.message.reply_text(
        f"Этап {stage_index}.\nВведите сумму этапа числом. Например: 150000 или 150000.50",
        reply_markup=cancel_markup(),
    )
    return CONTRACT_STAGE_AMOUNT


async def contract_stage_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or not update.message.text or update.effective_chat is None:
        return CONTRACT_STAGE_AMOUNT
    owner_chat_id, _ = owner_scope
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Введите сумму этапа или нажмите Отмена.", reply_markup=cancel_markup())
        return CONTRACT_STAGE_AMOUNT
    try:
        amount = parse_amount(raw)
    except ValueError:
        await update.message.reply_text("Не удалось распознать сумму. Пример: 150000.50")
        return CONTRACT_STAGE_AMOUNT

    context.user_data["contract_stage_items"].append(
        {
            "end_date": context.user_data["pending_contract_stage_end_date"],
            "amount": amount,
        }
    )
    total_count = context.user_data["contract_stage_count"]
    current_index = context.user_data["contract_stage_index"]
    if current_index < total_count:
        context.user_data["contract_stage_index"] = current_index + 1
        await update.message.reply_text(
            f"Этап {current_index + 1}.\nВведите дедлайн этапа в формате DD-MM-YYYY.",
            reply_markup=cancel_markup(),
        )
        return CONTRACT_STAGE_END_DATE

    stage_items = context.user_data["contract_stage_items"]
    contract_id = storage.add_contract(
        chat_id=owner_chat_id,
        title=context.user_data["contract_title"],
        description=context.user_data["contract_description"],
        end_date=stage_items[-1]["end_date"],
    )
    for index, item in enumerate(stage_items, start=1):
        storage.add_stage(
            contract_id=contract_id,
            position=index,
            notes="",
            end_date=item["end_date"],
            amount=item["amount"],
        )

    context.user_data.clear()
    await update.message.reply_text(
        f"Контракт сохранен с ID {contract_id}. Этапы созданы автоматически.",
        reply_markup=main_menu_markup(),
    )
    return ConversationHandler.END


async def start_stage_add_from_contract(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    storage: Storage = context.application.bot_data["storage"]
    chat = update.effective_chat
    owner_scope = await require_owner(update, context)
    if owner_scope is None or query is None or chat is None or query.data is None:
        return ConversationHandler.END
    owner_chat_id, _ = owner_scope
    await query.answer()
    try:
        contract_id = int(query.data.split(":", 1)[1])
    except (IndexError, ValueError):
        await query.edit_message_reply_markup(reply_markup=None)
        return ConversationHandler.END

    contract = storage.get_contract(owner_chat_id, contract_id)
    if contract is None:
        await query.edit_message_reply_markup(reply_markup=None)
        await query.message.reply_text("Контракт не найден.", reply_markup=main_menu_markup())
        return ConversationHandler.END

    context.user_data.clear()
    context.user_data["stage_contract_id"] = contract_id
    stages = storage.list_stages_for_contract(owner_chat_id, contract_id)
    await query.message.reply_text(
        f"Добавление этапа для контракта «{contract.title}».\n"
        f"Введите номер этапа. Например, следующий этап обычно {len(stages) + 1}.",
        reply_markup=cancel_markup(),
    )
    return STAGE_NAME


async def contract_end_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return ConversationHandler.END

    storage: Storage = context.application.bot_data["storage"]
    contract_id = storage.add_contract(
        chat_id=update.effective_chat.id,
        title=context.user_data["contract_title"],
        description=context.user_data["contract_description"],
        end_date=date.today(),
    )
    context.user_data.clear()
    await update.message.reply_text(
        f"Контракт сохранен с ID {contract_id}. Теперь можно добавить этапы через /new_stage.",
        reply_markup=main_menu_markup(),
    )
    return ConversationHandler.END


async def new_stage(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or update.effective_chat is None:
        return ConversationHandler.END
    owner_chat_id, _ = owner_scope
    context.user_data.clear()
    contracts = storage.list_contracts(owner_chat_id)
    if not contracts:
        await update.message.reply_text("Сначала добавьте контракт через /new_contract.", reply_markup=main_menu_markup())
        return ConversationHandler.END
    await update.message.reply_text(
        "Выберите контракт, к которому относится этап:",
        reply_markup=contract_selection_markup(contracts),
    )
    return STAGE_CONTRACT_ID


async def new_payment(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None:
        return ConversationHandler.END
    owner_chat_id, _ = owner_scope
    context.user_data.clear()
    contracts = storage.list_contracts(owner_chat_id)
    if not contracts:
        await update.message.reply_text("Сначала добавьте контракт через /new_contract.", reply_markup=main_menu_markup())
        return ConversationHandler.END
    await update.message.reply_text(
        "Выберите контракт, по которому поступила оплата:",
        reply_markup=contract_selection_markup(contracts),
    )
    return PAYMENT_CONTRACT_ID


async def stage_contract_id(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or not update.message.text or update.effective_chat is None:
        return STAGE_CONTRACT_ID
    owner_chat_id, _ = owner_scope
    raw = update.message.text.strip()
    if raw == MENU_CANCEL:
        return await cancel(update, context)
    try:
        contract_id = int(raw.split("|", 1)[0].strip())
    except ValueError:
        await update.message.reply_text("Выберите контракт кнопкой или введите его ID числом.")
        return STAGE_CONTRACT_ID

    contract = storage.get_contract(owner_chat_id, contract_id)
    if contract is None:
        await update.message.reply_text("Контракт не найден. Проверьте ID.")
        return STAGE_CONTRACT_ID

    context.user_data["stage_contract_id"] = contract_id
    stages = storage.list_stages_for_contract(owner_chat_id, contract_id)
    await update.message.reply_text(
        f"Введите номер этапа. Например, следующий этап обычно {len(stages) + 1}.",
        reply_markup=cancel_markup(),
    )
    return STAGE_NAME


async def stage_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message is None or not update.message.text:
        return STAGE_NAME
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Сначала введите номер этапа или нажмите Отмена.", reply_markup=cancel_markup())
        return STAGE_NAME
    try:
        position = int(raw)
    except ValueError:
        await update.message.reply_text("Номер этапа должен быть числом.", reply_markup=cancel_markup())
        return STAGE_NAME
    if position <= 0:
        await update.message.reply_text("Номер этапа должен быть больше нуля.", reply_markup=cancel_markup())
        return STAGE_NAME
    context.user_data["stage_position"] = position
    await update.message.reply_text("Введите примечание к этапу или отправьте '-' если не нужно.", reply_markup=cancel_markup())
    return STAGE_NOTES


async def stage_notes(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message is None or not update.message.text:
        return STAGE_NOTES
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Сначала введите примечание, '-' или нажмите Отмена.", reply_markup=cancel_markup())
        return STAGE_NOTES
    context.user_data["stage_notes"] = "" if raw == "-" else raw
    await update.message.reply_text("Введите дату окончания этапа в формате DD-MM-YYYY.", reply_markup=cancel_markup())
    return STAGE_END_DATE


async def stage_end_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or not update.message.text or update.effective_chat is None:
        return STAGE_END_DATE
    owner_chat_id, _ = owner_scope
    if is_menu_text(update.message.text.strip()):
        await update.message.reply_text("Введите дату этапа в формате DD-MM-YYYY или нажмите Отмена.", reply_markup=cancel_markup())
        return STAGE_END_DATE
    try:
        end_date = parse_date(update.message.text)
    except ValueError:
        await update.message.reply_text("Не удалось распознать дату. Нужен формат DD-MM-YYYY.")
        return STAGE_END_DATE

    contract = storage.get_contract(owner_chat_id, context.user_data["stage_contract_id"])
    if contract is None:
        context.user_data.clear()
        await update.message.reply_text("Контракт больше не найден. Повторите через /new_stage.")
        return ConversationHandler.END
    if end_date > contract.end_date:
        await update.message.reply_text(
            f"Этап не может заканчиваться позже контракта ({format_date(contract.end_date)}). Введите другую дату."
        )
        return STAGE_END_DATE
    context.user_data["stage_end_date"] = end_date
    await update.message.reply_text("Введите сумму этапа числом. Например: 150000 или 150000.50", reply_markup=cancel_markup())
    return STAGE_AMOUNT


async def stage_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or not update.message.text:
        return STAGE_AMOUNT
    if is_menu_text(update.message.text.strip()):
        await update.message.reply_text("Введите сумму этапа или нажмите Отмена.", reply_markup=cancel_markup())
        return STAGE_AMOUNT
    try:
        amount = parse_amount(update.message.text)
    except ValueError:
        await update.message.reply_text("Не удалось распознать сумму. Пример: 150000.50")
        return STAGE_AMOUNT

    stage_id = storage.add_stage(
        contract_id=context.user_data["stage_contract_id"],
        position=context.user_data["stage_position"],
        notes=context.user_data["stage_notes"],
        end_date=context.user_data["stage_end_date"],
        amount=amount,
    )
    context.user_data.clear()
    await update.message.reply_text(f"Этап сохранен с ID {stage_id}.", reply_markup=main_menu_markup())
    return ConversationHandler.END


async def payment_contract_id(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or not update.message.text:
        return PAYMENT_CONTRACT_ID
    owner_chat_id, _ = owner_scope
    raw = update.message.text.strip()
    if raw == MENU_CANCEL:
        return await cancel(update, context)
    try:
        contract_id = int(raw.split("|", 1)[0].strip())
    except ValueError:
        await update.message.reply_text("Выберите контракт кнопкой или введите его ID числом.")
        return PAYMENT_CONTRACT_ID
    contract = storage.get_contract(owner_chat_id, contract_id)
    if contract is None:
        await update.message.reply_text("Контракт не найден. Проверьте ID.")
        return PAYMENT_CONTRACT_ID
    context.user_data["payment_contract_id"] = contract_id
    await update.message.reply_text("Введите дату оплаты в формате DD-MM-YYYY.", reply_markup=cancel_markup())
    return PAYMENT_DATE


async def payment_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message is None or not update.message.text:
        return PAYMENT_DATE
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Введите дату оплаты в формате DD-MM-YYYY или нажмите Отмена.", reply_markup=cancel_markup())
        return PAYMENT_DATE
    try:
        parsed_date = parse_date(raw)
    except ValueError:
        await update.message.reply_text("Не удалось распознать дату. Нужен формат DD-MM-YYYY.")
        return PAYMENT_DATE
    context.user_data["payment_date"] = parsed_date
    await update.message.reply_text("Введите сумму оплаты числом. Например: 150000 или 150000.50", reply_markup=cancel_markup())
    return PAYMENT_AMOUNT


async def payment_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or not update.message.text:
        return PAYMENT_AMOUNT
    owner_chat_id, _ = owner_scope
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Введите сумму оплаты или нажмите Отмена.", reply_markup=cancel_markup())
        return PAYMENT_AMOUNT
    try:
        amount = parse_amount(raw)
    except ValueError:
        await update.message.reply_text("Не удалось распознать сумму. Пример: 150000.50")
        return PAYMENT_AMOUNT
    payment_id = storage.add_payment(
        owner_chat_id,
        context.user_data["payment_contract_id"],
        context.user_data["payment_date"],
        amount,
    )
    context.user_data.clear()
    await update.message.reply_text(
        f"Оплата сохранена с ID {payment_id}." if payment_id is not None else "Контракт не найден.",
        reply_markup=main_menu_markup(),
    )
    return ConversationHandler.END


async def edit_stage_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    storage: Storage = context.application.bot_data["storage"]
    chat = update.effective_chat
    owner_scope = await require_owner(update, context)
    if owner_scope is None or query is None or chat is None or query.data is None:
        return ConversationHandler.END
    owner_chat_id, _ = owner_scope
    await query.answer()
    try:
        stage_id = int(query.data.split(":", 1)[1])
    except (IndexError, ValueError):
        await query.edit_message_reply_markup(reply_markup=None)
        return ConversationHandler.END

    stage = storage.get_stage(owner_chat_id, stage_id)
    if stage is None:
        await query.edit_message_reply_markup(reply_markup=None)
        await query.message.reply_text("Этап не найден.", reply_markup=main_menu_markup())
        return ConversationHandler.END

    context.user_data["edit_stage_id"] = stage.id
    context.user_data["edit_contract_id"] = stage.contract_id
    context.user_data["edit_stage_name"] = stage.name
    context.user_data["edit_stage_notes"] = stage.notes
    context.user_data["edit_stage_end_date"] = stage.end_date
    context.user_data["edit_stage_amount"] = stage.amount

    await query.message.reply_text(
        f"Изменение {stage.name}.\nВведите новое примечание или '.' чтобы оставить текущее. Для пустого примечания введите '-'.",
        reply_markup=cancel_markup(),
    )
    return EDIT_STAGE_NOTES


async def edit_stage_notes(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message is None or not update.message.text:
        return EDIT_STAGE_NOTES
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Введите примечание, '.', '-' или нажмите Отмена.", reply_markup=cancel_markup())
        return EDIT_STAGE_NOTES
    if raw == "-":
        context.user_data["edit_stage_notes"] = ""
    elif raw != ".":
        context.user_data["edit_stage_notes"] = raw
    await update.message.reply_text("Введите новую дату этапа DD-MM-YYYY или '.' чтобы оставить текущую.", reply_markup=cancel_markup())
    return EDIT_STAGE_END_DATE


async def edit_stage_end_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or not update.message.text or update.effective_chat is None:
        return EDIT_STAGE_END_DATE
    owner_chat_id, _ = owner_scope
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Введите дату, '.' или нажмите Отмена.", reply_markup=cancel_markup())
        return EDIT_STAGE_END_DATE
    if raw != ".":
        try:
            end_date = parse_date(raw)
        except ValueError:
            await update.message.reply_text("Не удалось распознать дату. Нужен формат DD-MM-YYYY.")
            return EDIT_STAGE_END_DATE
        contract = storage.get_contract(owner_chat_id, context.user_data["edit_contract_id"])
        if contract is None:
            context.user_data.clear()
            await update.message.reply_text("Контракт больше не найден.", reply_markup=main_menu_markup())
            return ConversationHandler.END
        if end_date > contract.end_date:
            await update.message.reply_text(
                f"Этап не может заканчиваться позже контракта ({format_date(contract.end_date)}). Введите другую дату."
            )
            return EDIT_STAGE_END_DATE
        context.user_data["edit_stage_end_date"] = end_date
    await update.message.reply_text("Введите новую сумму или '.' чтобы оставить текущую.", reply_markup=cancel_markup())
    return EDIT_STAGE_AMOUNT


async def edit_stage_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or not update.message.text or update.effective_chat is None:
        return EDIT_STAGE_AMOUNT
    owner_chat_id, _ = owner_scope
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Введите сумму, '.' или нажмите Отмена.", reply_markup=cancel_markup())
        return EDIT_STAGE_AMOUNT
    if raw != ".":
        try:
            context.user_data["edit_stage_amount"] = parse_amount(raw)
        except ValueError:
            await update.message.reply_text("Не удалось распознать сумму. Пример: 150000.50")
            return EDIT_STAGE_AMOUNT

    updated = storage.update_stage(
        chat_id=owner_chat_id,
        stage_id=context.user_data["edit_stage_id"],
        name=context.user_data["edit_stage_name"],
        notes=context.user_data["edit_stage_notes"],
        end_date=context.user_data["edit_stage_end_date"],
        amount=context.user_data["edit_stage_amount"],
    )
    context.user_data.clear()
    await update.message.reply_text(
        "Этап изменен." if updated else "Этап не найден.",
        reply_markup=main_menu_markup(),
    )
    return ConversationHandler.END


async def start_payment_add_from_contract(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or query is None or query.data is None:
        return ConversationHandler.END
    owner_chat_id, _ = owner_scope
    await query.answer()
    try:
        contract_id = int(query.data.split(":", 1)[1])
    except (IndexError, ValueError):
        await query.edit_message_reply_markup(reply_markup=None)
        return ConversationHandler.END
    contract = storage.get_contract(owner_chat_id, contract_id)
    if contract is None:
        await query.message.reply_text("Контракт не найден.", reply_markup=main_menu_markup())
        return ConversationHandler.END
    context.user_data.clear()
    context.user_data["payment_contract_id"] = contract_id
    await query.message.reply_text(
        f"Добавление оплаты для контракта «{contract.title}».\nВведите дату оплаты в формате DD-MM-YYYY.",
        reply_markup=cancel_markup(),
    )
    return PAYMENT_DATE


async def show_contract_payments(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    storage: Storage = context.application.bot_data["storage"]
    scope = resolve_data_scope(update, context)
    if scope is None or query is None or query.data is None:
        return
    owner_chat_id, can_edit = scope
    await query.answer()
    try:
        contract_id = int(query.data.split(":", 1)[1])
    except (IndexError, ValueError):
        await query.edit_message_reply_markup(reply_markup=None)
        return
    contract = storage.get_contract(owner_chat_id, contract_id)
    if contract is None:
        await query.message.reply_text("Контракт не найден.", reply_markup=main_menu_markup() if can_edit else viewer_menu_markup())
        return
    payments = storage.list_payments_for_contract(owner_chat_id, contract_id)
    total_amount = sum(stage.amount for stage in storage.list_stages_for_contract(owner_chat_id, contract_id))
    paid_amount = sum(payment.amount for payment in payments)
    debt_amount = max(total_amount - paid_amount, 0.0)
    paid_percent = (paid_amount / total_amount * 100) if total_amount > 0 else 0.0
    lines = [f"Платежи по контракту «{escape(contract.title)}»:"]
    if payments:
        lines.append("")
        for payment in payments:
            lines.append(f"{format_date(payment.payment_date)} - {format_amount(payment.amount)}")
    else:
        lines.append("")
        lines.append("Платежей пока нет.")
    lines.append("")
    lines.append(f"Итого оплачено: {format_amount(paid_amount)} ({format_percent(paid_percent)})")
    lines.append(f"Общая сумма контракта: {format_amount(total_amount)}")
    lines.append(f"Долг: {format_amount(debt_amount)}")
    await query.message.reply_text(
        "\n".join(lines),
        parse_mode=ParseMode.HTML,
        reply_markup=payment_list_markup(contract_id, payments, can_edit),
    )


async def edit_payment_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or query is None or query.data is None:
        return ConversationHandler.END
    owner_chat_id, _ = owner_scope
    await query.answer()
    try:
        payment_id = int(query.data.split(":", 1)[1])
    except (IndexError, ValueError):
        await query.edit_message_reply_markup(reply_markup=None)
        return ConversationHandler.END
    payment = storage.get_payment(owner_chat_id, payment_id)
    if payment is None:
        await query.message.reply_text("Платеж не найден.", reply_markup=main_menu_markup())
        return ConversationHandler.END
    context.user_data["edit_payment_id"] = payment.id
    context.user_data["edit_payment_date"] = payment.payment_date
    context.user_data["edit_payment_amount"] = payment.amount
    await query.message.reply_text(
        f"Изменение платежа по контракту «{payment.contract_title}».\n"
        f"Введите новую дату DD-MM-YYYY или '.' чтобы оставить текущую ({format_date(payment.payment_date)}).",
        reply_markup=cancel_markup(),
    )
    return EDIT_PAYMENT_DATE


async def edit_payment_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message is None or not update.message.text:
        return EDIT_PAYMENT_DATE
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Введите дату, '.' или нажмите Отмена.", reply_markup=cancel_markup())
        return EDIT_PAYMENT_DATE
    if raw != ".":
        try:
            context.user_data["edit_payment_date"] = parse_date(raw)
        except ValueError:
            await update.message.reply_text("Не удалось распознать дату. Нужен формат DD-MM-YYYY.")
            return EDIT_PAYMENT_DATE
    await update.message.reply_text("Введите новую сумму или '.' чтобы оставить текущую.", reply_markup=cancel_markup())
    return EDIT_PAYMENT_AMOUNT


async def edit_payment_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or not update.message.text:
        return EDIT_PAYMENT_AMOUNT
    owner_chat_id, _ = owner_scope
    raw = update.message.text.strip()
    if is_menu_text(raw):
        await update.message.reply_text("Введите сумму, '.' или нажмите Отмена.", reply_markup=cancel_markup())
        return EDIT_PAYMENT_AMOUNT
    if raw != ".":
        try:
            context.user_data["edit_payment_amount"] = parse_amount(raw)
        except ValueError:
            await update.message.reply_text("Не удалось распознать сумму. Пример: 150000.50")
            return EDIT_PAYMENT_AMOUNT
    updated = storage.update_payment(
        owner_chat_id,
        context.user_data["edit_payment_id"],
        context.user_data["edit_payment_date"],
        context.user_data["edit_payment_amount"],
    )
    context.user_data.clear()
    await update.message.reply_text(
        "Платеж изменен." if updated else "Платеж не найден.",
        reply_markup=main_menu_markup(),
    )
    return ConversationHandler.END


async def delete_payment_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or query is None or query.data is None:
        return
    owner_chat_id, _ = owner_scope
    await query.answer()
    try:
        payment_id = int(query.data.split(":", 1)[1])
    except (IndexError, ValueError):
        await query.edit_message_reply_markup(reply_markup=None)
        return
    deleted = storage.delete_payment(owner_chat_id, payment_id)
    await query.edit_message_reply_markup(reply_markup=None)
    await query.message.reply_text(
        "Платеж удален." if deleted else "Платеж не найден.",
        reply_markup=main_menu_markup(),
    )


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    if update.message is not None:
        await update.message.reply_text("Текущий ввод отменен.", reply_markup=main_menu_markup())
    return ConversationHandler.END


async def delete_stage_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    storage: Storage = context.application.bot_data["storage"]
    chat = update.effective_chat
    owner_scope = await require_owner(update, context)
    if owner_scope is None or query is None or chat is None or query.data is None:
        return
    owner_chat_id, _ = owner_scope
    await query.answer()
    try:
        stage_id = int(query.data.split(":", 1)[1])
    except (IndexError, ValueError):
        await query.edit_message_reply_markup(reply_markup=None)
        return

    deleted = storage.delete_stage(owner_chat_id, stage_id)
    await query.edit_message_reply_markup(reply_markup=None)
    await query.message.reply_text(
        "Этап удален." if deleted else "Этап не найден.",
        reply_markup=main_menu_markup(),
    )


async def show_stage_status_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    storage: Storage = context.application.bot_data["storage"]
    chat = update.effective_chat
    owner_scope = await require_owner(update, context)
    if owner_scope is None or query is None or chat is None or query.data is None:
        return
    owner_chat_id, _ = owner_scope
    await query.answer()
    try:
        stage_id = int(query.data.split(":", 1)[1])
    except (IndexError, ValueError):
        await query.edit_message_reply_markup(reply_markup=None)
        return
    stage = storage.get_stage(owner_chat_id, stage_id)
    if stage is None:
        await query.message.reply_text("Этап не найден.", reply_markup=main_menu_markup())
        return
    await query.message.reply_text(
        f"Выберите статус для {stage.name}:",
        reply_markup=stage_status_markup(stage.id),
    )


async def set_stage_status(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    storage: Storage = context.application.bot_data["storage"]
    chat = update.effective_chat
    owner_scope = await require_owner(update, context)
    if owner_scope is None or query is None or chat is None or query.data is None:
        return
    owner_chat_id, _ = owner_scope
    await query.answer()
    payload = query.data.removeprefix(STAGE_STATUS_SET_PREFIX)
    try:
        stage_id_raw, status = payload.split(":", 1)
        stage_id = int(stage_id_raw)
    except ValueError:
        await query.edit_message_reply_markup(reply_markup=None)
        return
    if status not in STAGE_STATUSES:
        await query.edit_message_reply_markup(reply_markup=None)
        return
    updated = storage.update_stage_status(owner_chat_id, stage_id, status)
    await query.edit_message_reply_markup(reply_markup=None)
    await query.message.reply_text(
        f"Статус обновлен: {status_emoji(status)} {status_label(status)}" if updated else "Этап не найден.",
        reply_markup=main_menu_markup(),
    )


async def contract_manage_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    storage: Storage = context.application.bot_data["storage"]
    chat = update.effective_chat
    owner_scope = await require_owner(update, context)
    if owner_scope is None or query is None or chat is None or query.data is None:
        return
    owner_chat_id, _ = owner_scope
    await query.answer()
    try:
        contract_id = int(query.data.split(":", 1)[1])
    except (IndexError, ValueError):
        await query.edit_message_reply_markup(reply_markup=None)
        return

    contract = storage.get_contract(owner_chat_id, contract_id)
    if contract is None:
        await query.message.reply_text("Контракт не найден.", reply_markup=main_menu_markup())
        return

    stages = storage.list_stages_for_contract(owner_chat_id, contract_id)
    await query.message.reply_text(
        f"Управление этапами контракта «{contract.title}»:",
        reply_markup=contract_manage_actions_markup(contract_id),
    )
    if not stages:
        await query.message.reply_text("Этапов пока нет.")
        return

    for stage in stages:
        await query.message.reply_text(
            f"{status_emoji(stage.status)} Этап {stage.position} - {format_amount(stage.amount)} - {format_date(stage.end_date)}",
            reply_markup=stage_actions_markup(stage.id),
        )


async def list_contracts(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    storage: Storage = context.application.bot_data["storage"]
    scope = resolve_data_scope(update, context)
    if scope is None or update.message is None or update.effective_chat is None:
        return
    owner_chat_id, can_edit = scope

    contracts = storage.list_contracts(owner_chat_id)
    if not contracts:
        await update.message.reply_text("Контрактов пока нет.", reply_markup=main_menu_markup() if can_edit else viewer_menu_markup())
        return

    for contract in contracts:
        stages = storage.list_stages_for_contract(owner_chat_id, contract.id)
        contract.paid_amount = storage.contract_payment_total(owner_chat_id, contract.id)
        await update.message.reply_text(
            render_contract_with_stages(contract, stages),
            parse_mode=ParseMode.HTML,
            reply_markup=contract_actions_markup(contract.id, can_edit),
        )
    await update.message.reply_text("Список показан.", reply_markup=main_menu_markup() if can_edit else viewer_menu_markup())


async def upcoming(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    storage: Storage = context.application.bot_data["storage"]
    scope = resolve_data_scope(update, context)
    if scope is None or update.message is None or update.effective_chat is None:
        return
    owner_chat_id, can_edit = scope
    days = 30
    if context.args:
        try:
            days = max(1, int(context.args[0]))
        except ValueError:
            await update.message.reply_text("Если указывать число, то в формате /upcoming 30")
            return

    items = storage.upcoming_items(owner_chat_id, within_days=days)
    if not items:
        nearest = storage.nearest_item(owner_chat_id)
        if nearest is None:
            await update.message.reply_text(f"На ближайшие {days} дней дедлайнов нет.", reply_markup=main_menu_markup() if can_edit else viewer_menu_markup())
            return
        if nearest["entity_type"] == "stage":
            stage = storage.get_stage(owner_chat_id, nearest["entity_id"])
            if stage is not None:
                details = (
                    "Самый ближайший известный дедлайн:\n"
                    f"Контракт: {stage.contract_title}\n"
                    f"Этап: {stage.name}\n"
                    f"Сумма: {format_amount(stage.amount)}\n"
                    f"Дедлайн: {format_date(nearest['end_date'])}\n"
                    f"Осталось: {nearest['days_left']} дн."
                )
            else:
                details = (
                    "Самый ближайший известный дедлайн:\n"
                    f"Контракт: не определен\n"
                    f"Этап: {nearest['title']}\n"
                    f"Дедлайн: {format_date(nearest['end_date'])}\n"
                    f"Осталось: {nearest['days_left']} дн."
                )
        else:
            contract = storage.get_contract(owner_chat_id, nearest["entity_id"])
            contract_amount = 0.0
            if contract is not None:
                contract_amount = sum(stage.amount for stage in storage.list_stages_for_contract(owner_chat_id, contract.id))
            details = (
                "Самый ближайший известный дедлайн:\n"
                f"Контракт: {nearest['title']}\n"
                f"Этап: нет\n"
                f"Сумма: {format_amount(contract_amount)}\n"
                f"Дедлайн: {format_date(nearest['end_date'])}\n"
                f"Осталось: {nearest['days_left']} дн."
            )
        await update.message.reply_text(
            f"На ближайшие {days} дней дедлайнов нет.\n"
            f"{details}",
            reply_markup=main_menu_markup() if can_edit else viewer_menu_markup(),
        )
        return

    lines = [f"Ближайшие дедлайны на {days} дней:"]
    for item in items:
        kind = "Контракт" if item["entity_type"] == "contract" else "Этап"
        suffix = ""
        if item["entity_type"] == "stage":
            stage = storage.get_stage(owner_chat_id, item["entity_id"])
            if stage is not None:
                suffix = f" | {status_emoji(stage.status)} {status_label(stage.status)} | сумма {format_amount(stage.amount)}"
        lines.append(
            f"{kind} #{item['entity_id']}: {item['title']} | {format_date(item['end_date'])} | {format_due(item['days_left'])}{suffix}"
        )
    await update.message.reply_text("\n".join(lines), reply_markup=main_menu_markup() if can_edit else viewer_menu_markup())


async def delete_contract(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or update.effective_chat is None:
        return
    owner_chat_id, _ = owner_scope
    if not context.args:
        await update.message.reply_text("Формат: /delete_contract ID", reply_markup=main_menu_markup())
        return
    try:
        contract_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("ID должен быть числом.")
        return
    deleted = storage.delete_contract(owner_chat_id, contract_id)
    await update.message.reply_text(
        "Контракт удален." if deleted else "Контракт не найден.",
        reply_markup=main_menu_markup(),
    )


async def delete_stage(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    storage: Storage = context.application.bot_data["storage"]
    owner_scope = await require_owner(update, context)
    if owner_scope is None or update.message is None or update.effective_chat is None:
        return
    owner_chat_id, _ = owner_scope
    if not context.args:
        await update.message.reply_text("Формат: /delete_stage ID", reply_markup=main_menu_markup())
        return
    try:
        stage_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("ID должен быть числом.")
        return
    deleted = storage.delete_stage(owner_chat_id, stage_id)
    await update.message.reply_text(
        "Этап удален." if deleted else "Этап не найден.",
        reply_markup=main_menu_markup(),
    )


async def menu_new_contract(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await new_contract(update, context)


async def menu_new_stage(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await new_stage(update, context)


async def menu_new_payment(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await new_payment(update, context)


async def menu_contracts(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await list_contracts(update, context)


async def menu_export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await export_excel(update, context)


async def menu_upcoming(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await upcoming(update, context)


async def menu_invite(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await invite_viewer(update, context)


async def menu_access_list(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await access_list(update, context)


async def send_reminders(context: ContextTypes.DEFAULT_TYPE) -> None:
    storage: Storage = context.application.bot_data["storage"]
    candidates = storage.reminder_candidates(REMINDER_DAYS)
    for item in candidates:
        kind = "контракт" if item["entity_type"] == "contract" else "этап"
        text = (
            f"Напоминание: скоро заканчивается {kind}.\n"
            f"{item['title']}\n"
            f"Дата окончания: {format_date(item['end_date'])}\n"
            f"Осталось: {item['days_left']} дн."
        )
        recipients = [item["chat_id"], *storage.viewers_for_owner(item["chat_id"])]
        for recipient_chat_id in recipients:
            if storage.reminder_already_sent(
                recipient_chat_id,
                item["entity_type"],
                item["entity_id"],
                item["days_left"],
                item["end_date"],
            ):
                continue
            try:
                await context.bot.send_message(chat_id=recipient_chat_id, text=text)
                storage.mark_reminder_sent(
                    recipient_chat_id,
                    item["entity_type"],
                    item["entity_id"],
                    item["days_left"],
                    item["end_date"],
                )
            except Exception:
                LOGGER.exception("Failed to send reminder to chat_id=%s", recipient_chat_id)


def build_application() -> Application:
    load_dotenv()
    token = os.environ["BOT_TOKEN"]
    timezone = ZoneInfo(os.getenv("BOT_TIMEZONE", "Asia/Vladivostok"))
    db_path = os.getenv("DB_PATH", "contracts.db")

    app = Application.builder().token(token).build()
    app.bot_data["storage"] = Storage(db_path)

    contract_conv = ConversationHandler(
        entry_points=[
            CommandHandler("new_contract", new_contract),
            MessageHandler(filters.Regex(f"^{MENU_NEW_CONTRACT}$"), menu_new_contract),
        ],
        states={
            CONTRACT_TITLE: [MessageHandler(filters.TEXT & ~filters.COMMAND, contract_title)],
            CONTRACT_DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, contract_description)],
            CONTRACT_STAGE_COUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, contract_stage_count)],
            CONTRACT_STAGE_END_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, contract_stage_end_date)],
            CONTRACT_STAGE_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, contract_stage_amount)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        map_to_parent=None,
    )

    stage_conv = ConversationHandler(
        entry_points=[
            CommandHandler("new_stage", new_stage),
            MessageHandler(filters.Regex(f"^{MENU_NEW_STAGE}$"), menu_new_stage),
        ],
        states={
            STAGE_CONTRACT_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, stage_contract_id)],
            STAGE_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, stage_name)],
            STAGE_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, stage_notes)],
            STAGE_END_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, stage_end_date)],
            STAGE_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, stage_amount)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    edit_stage_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(edit_stage_start, pattern=f"^{STAGE_EDIT_PREFIX}")],
        states={
            EDIT_STAGE_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_stage_notes)],
            EDIT_STAGE_END_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_stage_end_date)],
            EDIT_STAGE_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_stage_amount)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    payment_conv = ConversationHandler(
        entry_points=[
            CommandHandler("new_payment", new_payment),
            MessageHandler(filters.Regex(f"^{MENU_NEW_PAYMENT}$"), menu_new_payment),
            CallbackQueryHandler(start_payment_add_from_contract, pattern=f"^{CONTRACT_ADD_PAYMENT_PREFIX}"),
        ],
        states={
            PAYMENT_CONTRACT_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, payment_contract_id)],
            PAYMENT_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, payment_date)],
            PAYMENT_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, payment_amount)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    edit_payment_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(edit_payment_start, pattern=f"^{PAYMENT_EDIT_PREFIX}")],
        states={
            EDIT_PAYMENT_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_payment_date)],
            EDIT_PAYMENT_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_payment_amount)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(MessageHandler(filters.Regex(f"^{MENU_CANCEL}$"), cancel))

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("invite_viewer", invite_viewer))
    app.add_handler(CommandHandler("access_list", access_list))
    app.add_handler(CommandHandler("export_excel", export_excel))
    app.add_handler(CommandHandler("contracts", list_contracts))
    app.add_handler(CommandHandler("upcoming", upcoming))
    app.add_handler(CommandHandler("delete_contract", delete_contract))
    app.add_handler(CommandHandler("delete_stage", delete_stage))
    app.add_handler(CallbackQueryHandler(contract_manage_menu, pattern=f"^{CONTRACT_MANAGE_PREFIX}"))
    app.add_handler(CallbackQueryHandler(show_contract_payments, pattern=f"^{CONTRACT_PAYMENTS_PREFIX}"))
    app.add_handler(CallbackQueryHandler(start_stage_add_from_contract, pattern=f"^{CONTRACT_ADD_STAGE_PREFIX}"))
    app.add_handler(CallbackQueryHandler(delete_payment_callback, pattern=f"^{PAYMENT_DELETE_PREFIX}"))
    app.add_handler(CallbackQueryHandler(revoke_access, pattern=f"^{ACCESS_REVOKE_PREFIX}"))
    app.add_handler(CallbackQueryHandler(show_stage_status_menu, pattern=f"^{STAGE_STATUS_PREFIX}"))
    app.add_handler(CallbackQueryHandler(set_stage_status, pattern=f"^{STAGE_STATUS_SET_PREFIX}"))
    app.add_handler(CallbackQueryHandler(delete_stage_callback, pattern=f"^{STAGE_DELETE_PREFIX}"))
    app.add_handler(MessageHandler(filters.Regex(f"^{MENU_CONTRACTS}$"), menu_contracts))
    app.add_handler(MessageHandler(filters.Regex(f"^{MENU_EXPORT_EXCEL}$"), menu_export_excel))
    app.add_handler(MessageHandler(filters.Regex(f"^{MENU_UPCOMING}$"), menu_upcoming))
    app.add_handler(MessageHandler(filters.Regex(f"^{MENU_INVITE}$"), menu_invite))
    app.add_handler(MessageHandler(filters.Regex(f"^{MENU_ACCESS_LIST}$"), menu_access_list))
    app.add_handler(contract_conv)
    app.add_handler(stage_conv)
    app.add_handler(payment_conv)
    app.add_handler(edit_stage_conv)
    app.add_handler(edit_payment_conv)

    app.job_queue.run_daily(
        send_reminders,
        time=time(hour=9, minute=0, tzinfo=timezone),
        name="deadline-reminders",
    )
    return app


def main() -> None:
    application = build_application()
    LOGGER.info("Bot is starting")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
